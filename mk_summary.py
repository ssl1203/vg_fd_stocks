#Testing 2023-07-14 AM
# edited by Windows Chrome APP 
#Testing Git 2023-07-06
import csv
import re
from numpy import float32
#from pywin.dialogs.list import test
import xlsxwriter
import pandas as pd
import yfinance as yf # backup for mutual fund

import xlwings
import shutil
from datetime import date
from datetime import datetime
from datetime import time

import os
import platform

# install lib
# >pip install requests
# >pip install beautifulsoup4
# >pip install lxml
from bs4 import BeautifulSoup
import requests
import lxml


if (platform.system()=='Darwin'):
    g_data_path = f'{os.path.dirname(__file__)}/../vg_fd_stocks_data/'
else: 
    g_data_path = f"{os.getcwd()}\\..\\vg_fd_stocks_data\\" 


g_yf_export_vg_fd = f'{g_data_path}yf_export_vg_fd.csv'
g_xl_summary_file = f'{g_data_path}summary-{str(date.today())}.xlsx'
g_yf_export_tw = f'{g_data_path}yf_export_tw.csv'
global g_fd_mephy_download_csv
global g_fd_download_csv
global g_vg_download_csv


def g_init():
    #global g_data_path
    #global g_yf_export_vg_fd
    global g_fd_mephy_download_csv
    global g_fd_download_csv
    global g_vg_download_csv

    if (not os.path.exists(g_data_path)):
        print(f'***** Fatal error, expected fold not exist ({g_data_path})')
        exit()   

    if (not os.path.exists(g_xl_summary_file)):
        wb = xw.Book()
        #wb.sheets.add('Sheet1')
        wb.sheets.add('Sheet2')
        wb.save(g_xl_summary_file)
        #wb.save("123.")
        wb.close()

    #g_yf_export_vg_fd = f'{g_data_path}yf_export_vg_fd.csv'

    # g_fd_mephy_download_csv = f"{g_data_path}fd_mephy.csv"
    # if not os.path.exists(g_fd_mephy_download_csv):
    #     print(f'***** Fatal error, expected file not exist ({g_fd_mephy_download_csv}')
    #     exit()
        

    g_fd_download_csv = f"{g_data_path}fd.csv"
    if not os.path.exists(g_fd_download_csv):
        print(f'***** Fatal error, expected file not exist ({g_fd_download_csv}')
        exit()

    g_vg_download_csv = f"{g_data_path}vg.csv"
    if not os.path.exists(g_vg_download_csv):
        print(f'***** Fatal error, expected file not exist ({g_vg_download_csv}')
        exit()


import xlwings as xw
import numpy as np

import os
#so.getcwd()
import openpyxl


g_row_account = 1
g_row_sum = 2

g_first_data_row = 3

g_first_data_col = 'E'
g_last_data_col = 'E'   # = chr(ord(g_first_data_col) + number_of_accounts - 1)

vg_fd_sheet_name = 'vg+fd'
tw_sheet_name = 'tw'

#yahoo_fin
#http://theautomatic.net/2018/07/31/how-to-get-live-stock-prices-with-python/
#http://theautomatic.net/yahoo_fin-documentation/
#http://theautomatic.net/about-me/
from yahoo_fin import stock_info as si

#si.get_quote_table('NVDA')['PE Ratio (TTM)'] 
#http://theautomatic.net/yahoo_fin-documentation/#stock_info

#  hp_laptop path
#g_working_path = 'C:/Users/seans/OneDrive/80-股市/000-download/2021-07-31/' 


g_vg_dict = {"60965808": "VG_Sean_IRA",
        "58513150":"VG_Sean_ROTH",
        "31141191": "VG_JOIN", 
        "52028985": "VG_Mephy_IRA",
        "15094632":"VG_Mephy_ROTH"}


g_fd_dict = {
        "X38745588" : "FD_JOIN", 
        "168682748":"FD_Sean_IRA",
        "229713633":"FD_Sean_ROTH",
        "32330":"FD_Sean_IRA",
        "225995814" : "FD_Mephy_IRA",
        "233641339":"FD_Mephy_HD"}

# g_fd_mephy_dict = { "225995814" : "FD_Mephy_IRA", 
#         "233641339":"FD_Mephy_HD"}




#symb_dict = {'GOOGL': 'GOOG','SPY':'VOO','VTI':'VOO','FZROX':'VOO'}
#value_dict = {}

col_header_vg_fd = [
    "Symbol",
    "Price",
    "Total Value",
    "Total Shares",
    "VG_Sean_IRA",     
    "VG_Sean_ROTH",
    "VG_JOIN", 
    "VG_Mephy_IRA" ,  
    "VG_Mephy_ROTH" ,

    "FD_JOIN"    ,    
    "FD_Sean_IRA" ,
    "FD_Sean_ROTH" ,  

    "FD_Mephy_IRA", 
    "FD_Mephy_HD",
]

col_header_tw = [
    "Symbol",
    "Price",
    "Total Value",
    "Total Shares",
    "yt-sean",     
    "yt-nathan",
    "yt-mom", 
    "cathay" ,  
    "mega" 
]

def get_current_stock_price(symb):

    if symb == 'USD=X' or symb=='CASH' or symb=='$$CASH':
        return 1
    
    price = si.get_live_price(symb) 

    #get last close price if live_price not available 
    if pd.isna(price) :
        try:
            stock_info = yf.Ticker(symb)
            data = stock_info.history(period="1wk") # '1mo' '20mo' 
            prices = data['Close']
            price = prices.iloc[-1]
            print(f'Got last close : {symb}')
        except:
            price=0
    return round(price ,2)   

#########################################
######### Get ETF Top Holdings ##########
######################################### 
mega_8 = ['AAPL', 'MSFT', 'AMZN', 'NVDA', 'META', 'TSLA', 'GOOGL', 'GOOG', 'TSM']
mega8_value_dict = {'AAPL':0, 'MSFT':0, 'AMZN':0, 'NVDA':0, 'META':0, 'TSLA':0, 'GOOGL':0, 'GOOG':0, 'TSM':0}

def GetHoldings(etf_ticker):
    crawler_headers = {
      'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'
    }
    etf_html = requests.get(f"https://finance.yahoo.com/quote/{etf_ticker}/holdings?p={etf_ticker}",headers=crawler_headers)
    top_10 = []
    etf_page_content = BeautifulSoup(etf_html.content,'lxml')

    #find the top holdings table
    the_table = etf_page_content.find('table', {'class': 'W(100%) M(0) BdB Bdc($seperatorColor)'})
    if the_table==None:
        return []

    top_10_dict = {}
    top_10_table = the_table.find_all('tr',   {'class': 'Ta(end) BdT Bdc($seperatorColor) H(36px)'})
    for item in top_10_table: 
        symb = item.find('a',{'class','Fz(s) Ell Fw(b) C($linkColor)'}).text
        percent = item.find('td',{'class',''}).text
        top_10.append([f'{symb}',float(percent.replace('%',''))])
        top_10_dict[f'{symb}'] = float(percent.replace('%',''))
    return top_10_dict
qqq_top_10_dict = {}
def calc_mega_8_holdings(df_all_stocks):
    for index, row in df_all_stocks.iterrows():
        symb = row[0]
        value = float(row[2])
        if symb  in mega_8 :
            mega8_value_dict[f'{symb}'] += value
            continue
        top_10_dict = GetHoldings(symb)
        if len(top_10_dict) == 0:
            continue
        for item in top_10_dict:
            if item in mega_8:
                #print(f' add etf ({symb}) to {item} ', float(top_10_dict[f'{item}'])/100*value)                 
                mega8_value_dict[f'{item}'] += float(top_10_dict[f'{item}'])/100*value
                if symb=='QQQ':
                    qqq_top_10_dict[f'{item}'] = float(top_10_dict[f'{item}'])/100
    #consolidate GOOGL to GOOG
    mega8_value_dict['GOOG'] += mega8_value_dict['GOOGL']
    del mega8_value_dict['GOOGL']

    if qqq_top_10_dict['GOOG'] > 0 and qqq_top_10_dict['GOOGL'] > 0:
        qqq_top_10_dict['GOOG']+=qqq_top_10_dict['GOOGL']
        del qqq_top_10_dict['GOOGL']

    #quick fix to include 2330.tw to TSM
    mega8_value_dict['TSM'] += 150000


def add_shares(matrix,acc_name,new_symb,shares,price,col_header,symb_list):
    
    try:
        idx_col = col_header.index(acc_name)
    except:
        print(f'*** Error : account {acc_name} not found')
        return False

    try:
        #fidility may have duplicate cash positions COREXX and SPAXX
        if new_symb not in symb_list:
            symb_list.append(new_symb)

            #init all cells in new row to 0's
            new_row = [0 for x in range(len(col_header))]
            new_row[0] = new_symb
            new_row[1] = price
            matrix.append(new_row)
        
        idx_row = symb_list.index(new_symb)
        
        #update the total value of a stock in the given accont
        #idx_row = idx of symbol, idx_col = idx of account
        matrix[idx_row][idx_col] = shares+matrix[idx_row][idx_col]  #account
    #     if new_symb == "$$CASH":
    #         price = 1
    #     else :
    #         price = get_current_stock_price(new_symb)
    #     matrix[idx_row][col_header.index("Price")]=price
    except:
        print(f"*** Fatal Error (159) - AddShare {acc_name} {new_symb} failed")   
def vanguard_reader(matrix,accounts_dict,csv_file_name,col_header,symb_list):
    
    count_0 = 0
    count_line = 0

    try:
        csv_file  = open(csv_file_name)
        csv_reader = csv.reader(csv_file, delimiter=',')  
    except:
        print(f'**** Failed to open {csv_file_name}')
        return False
    
    for row in csv_reader:
        count_line+=1
        if (count_line==1):  # skip header line
            continue
        if len(row)==0 or row[0]=='':      # terminate if two consecutive null lines
            count_0+=1
            if (count_0>1):
                break 
            continue
        count_0=0        
        acc_name = accounts_dict.get(row[0],'undefined')
        if acc_name=='undefined':
            print(f"Invalid account number {row[0]}")
            continue
        
        try :
            symb = row[2]
            shares = float(row[3])
            price = float(row[4])
            
            if int(price) == 1:
                symb = '$$CASH'
            add_shares(matrix,acc_name,symb,shares,price,col_header,symb_list)
        except :
            print(f'*** Fatal Error (202) : Not able save {row}')
            return False                         
    return True
def fidelity_reader(matrix,accounts_dict,csv_file_name,col_header,symb_list):

    fd_last_acc = ''
    try:
        count_0 = 0
        with open(csv_file_name) as csv_file:
            #print(f"File Opened : [{csv_file_name}]")
            fd_csv_reader = csv.reader(csv_file, delimiter=',')
            for fd_row in fd_csv_reader:
                count_0+=1
                if (count_0 ==1):  # skip the first line only
                    continue
                if len(fd_row)==0 or fd_row[0]=='':
                    break

                fd_acc_name = accounts_dict.get(fd_row[0],'undefined')
                if fd_acc_name== 'undefined' :
                    if (fd_row[0]!='X38745588'):
                        print(f"*** Warning (215) : account : ({fd_row[0]}  Sumb : ({fd_row[2]})) not processed")
                    continue

                if (fd_row[2]=='656568508'):
                    continue                         

                try :
                    if (fd_row[2]=='Pending Activity'):
                        shares = float(re.sub('[$,]', '', fd_row[7]))
                        add_shares(matrix,fd_acc_name,'$$CASH',shares,1,col_header,symb_list) 
                    else:
                        symb = fd_row[2]
                        price=0
                        share=0

                        price_str = re.sub('[$,]', '', fd_row[5])  #use substitue function to remove '$' and ','
                        price = float(price_str)    
                        share_str = re.sub('[,]', '', fd_row[4])
                        shares = float(share_str)
                        
                        #save_to_holding_pool(symb,shares,price,fd_acc_name,fd_row[0])
                        if int(price) == 1:
                            symb = '$$CASH' 
                        add_shares(matrix,fd_acc_name,symb,shares,price,col_header,symb_list)          
           
                except ValueError:
                    print(f'***Fatal Error (242) : add_shares symb={symb} failed') 

    except:
        print(f'***Fatal Error (246) : Unable process [{csv_file_name}]') 
        

    return


def update_summary_sheet():
    print("Creating summary sheet")

    sheet_name = "Summary"
#########################################
    try :
        wb = xw.Book(g_xl_summary_file)
    except:
        print(f'**** Failed to open {g_xl_summary_file} (302)')
        return

    try:
        wb.sheets.add(sheet_name)
        ws_summary = wb.sheets[sheet_name]
        #ws_summary.range("A1:E25").clear_contains()
    except:
        #wb.sheets.add(sheet_name)
        ws_summary = wb.sheets[sheet_name]

    ws_summary.range('A1:A30').font.bold = True
    ws_summary.range('D1:D30').font.bold = True
    ws_summary.range('E1:E1').font.bold = True

    
    # for border_id in rng(7,13):
    #     rng.Borders(border_id).LineStyle=1
    #     rng.Borders(border_id).Weight=2



####################################
   

    ws_vg_fd = wb.sheets[vg_fd_sheet_name]
    ws_vg_fd.range('A1:Z1').font.bold = True
    ws_vg_fd.range('A1:A30').font.bold = True


    ws_tw = wb.sheets[tw_sheet_name]

###############################################
    array_2d_vg = pd.DataFrame(ws_vg_fd.range('E1:N2').value)

    ws_summary["A6"].options(pd.DataFrame, header=0, index=False,transpose=True).value = array_2d_vg

    ######### update indices ###########
    #w_sheet.write('B2','S&P-500')
    #w_sheet.write('C2',si.get_live_price('^GSPC'))

    ntd_x_rate = get_current_stock_price("USDTWD=X")
    chf_x_rate = get_current_stock_price("CHFUSD=X")


    ws_summary['B1'].value=f'{date.today()}'
    ws_summary['B2'].value=get_current_stock_price("^GSPC")
    ws_summary['B3'].value=get_current_stock_price("^IXIC")
    ws_summary['B4'].value=get_current_stock_price("0050.TW")
    ws_summary['B5'].value=get_current_stock_price("QQQ")



    ws_summary['A1'].value='Date'
    ws_summary['A2'].value='S&P'
    ws_summary['A3'].value='Nasdaq'
    ws_summary['A4'].value='TW0050'
    ws_summary['A5'].value='QQQ'    

    


    ws_summary.range(f'B2:B19').number_format =    "#,##0"
    ws_summary.range(f'E1:E13').number_format =    "#,##0"
    ws_summary.range(f'B20:B21').number_format =   "##.00"

###############################################################
    ws_summary['D1'].value='Total'
    ws_summary['E1'].value='=SUM(B6:B18)'

    ws_summary['D2'].value='VG total'
    ws_summary['E2'].value = '=SUM(B6:B10)'

    ws_summary['D3'].value='Fidelity'
    ws_summary['E3'].value = '=SUM(B11:B15)'

    #ws_summary['D4'].value='FD Mephy'
    #ws_summary['E4'].value='=SUM(B15:B16)+B12'
###################################################
    ws_summary['D6'].value='Sean_IRA'
    ws_summary['E6'].value='=B6+B12'

    ws_summary['D7'].value='Mephy_IRA'
    ws_summary['E7'].value='=B9+B14'

    ws_summary['D8'].value='Sean_Roth'
    ws_summary['E8'].value='=B7+B13'

    ws_summary['D9'].value='Mpehy_Roth'
    ws_summary['E9'].value='=B10+B15'

    ws_summary['D10'].value='Saving'
    ws_summary['E10'].value='=B8+B11+B16+B17+B18'
##################################################


    ws_summary['A16'].value = 'QQQ192'
    ws_summary['B16'].value = get_current_stock_price("QQQ")*192

    ws_summary['A17'].value='Taiwan'

    ws_summary['B17'].value =  ws_tw['C2'].value/ntd_x_rate

    ws_summary['A18'].value='Tecan Stock'
    total_tecn = (get_current_stock_price("TECN.SW")*490-(264*228.6)-(226*236))
    ws_summary['B18'].value = total_tecn*chf_x_rate


    ws_summary['A20'].value='USDTWD=X'
    ws_summary['B20'].value = ntd_x_rate
    ws_summary['A21'].value='CHFUSD=X'
    ws_summary['B21'].value = chf_x_rate

##################################################################
############  Gega_8 #############################################
##############################################################
    ws_summary['D12'].value =    'Mega_8 %'
    ws_summary['G12'].value =    'QQQ %'
    ws_summary.range(f'F10:F25').number_format =   "##.00"
    ws_summary.range(f'F13:F20').number_format =    '0.00%'
    ws_summary.range(f'G13:G20').number_format =    '0.00%'
    ws_summary.range(f'E13:E20').number_format =    "$#,##0"

    ws_summary['F13'].value =    '=E13/E1'
    ws_summary['F14'].value =    '=E14/E1'
    ws_summary['F15'].value =    '=E15/E1'
    ws_summary['F16'].value =    '=E16/E1'
    ws_summary['F17'].value =    '=E17/E1'
    ws_summary['F18'].value =    '=E18/E1'
    ws_summary['F19'].value =    '=E19/E1'
    ws_summary['F20'].value =    '=E20/E1'
    global mega8_value_dict
    global qqq_top_10_dict
    col_id = ord('D')
    row_id = 13
    for row in mega8_value_dict:
        print(f'mega_8 {row} =' , round(mega8_value_dict[f'{row}'],2))
        ws_summary[f'{chr(col_id)}{row_id}'].value = row
        ws_summary[f'{chr(col_id+1)}{row_id}'].value = round(mega8_value_dict[f'{row}'],2)

        if row != 'TSM':
            ws_summary[f'{chr(col_id+3)}{row_id}'].value = round(qqq_top_10_dict[f'{row}'],2)

        row_id+=1
##################################################################


    print('----------- Summary Sheet Complete ---------------')
    #ws_summary['E6'].formula = tecn_total * ws_summary['E16'].value
   







########################################################################
def save2excel(sheet_name,df_sorted,col_header):

    first_acc_col = 'E'
    last_acc_col = chr(ord(first_acc_col)+df_sorted.shape[1]-4-1)  # 4 is the none account columns (overhead columns)
    first_stock_row = 3
    last_stock_row = first_stock_row+df_sorted.shape[0]-1

    try :     
        wb = xw.Book(g_xl_summary_file)
    except:
        print(f'**** Failed to open {g_xl_summary_file}')
        return
    
######################################
# Cleanup vg_fd by delete it and create a fresh one
    try:
        if (not sheet_name in wb.sheet_names):
            wb.sheets.add(sheet_name)
    except:
        print('**** ERROR 458 *****', sheet_name)
####################################
    ws = wb.sheets[f'{sheet_name}']

    #df = pd.DataFrame(data_arry)
    ws[f"A1"].options(pd.DataFrame, header=0, index=False, transpose=True).value = pd.DataFrame(col_header)
    ws[f"A{first_stock_row}"].options(pd.DataFrame, header=0, index=False, expand='table').value = df_sorted

    ws[f'C2'].formula=f'=SUM(C{first_stock_row}:C{last_stock_row})' 

    for col_id in range(ord(first_acc_col), ord(last_acc_col)+1):
        ws[f'{chr(col_id)}2'].formula = f'=SUMPRODUCT(B{first_stock_row}:B{last_stock_row},{chr(col_id)}{first_stock_row}:{chr(col_id)}{last_stock_row})' 

    ws['A2'].formula = f'=TEXT(TODAY(), "yyyy-mm-dd")'
    ws['B2'].value = datetime.now().strftime("%X")

    
    ws.range(f'B3:{last_acc_col}{last_stock_row}').number_format =    "#,##0"
    ws.range(f'C2:{last_acc_col}2').number_format =    "#,##0"
    ws.range('C2:C2').font.bold = True
    #ws_vg_fd.range(f'C:{g_row_sum}').number_format = "#,##0.#000"
    print(f'================ Total Value :',  "{:,.0f}".format(ws[f'C2'].value), ' ===================')

    return True


########################################################################
def make_export_for_yf(df_holdings,yf_export_file):
    title = 'Symbol,Current Price,Date,Time,Change,Open,High,Low,Volume,Trade Date,Purchase Price,Quantity,Commission,High Limit,Low Limit,Comment\r'
    
    try:
        with open(yf_export_file, 'w') as yf_csv:
            yf_csv.write(title)
            # reverse order from small to large position
            for stock_position in df_holdings.iloc[::-1].itertuples(): 
                yf_csv.write(f'{stock_position[1]},{stock_position[2]},,,,,,,,,,{stock_position[4]},,,,\r') 
            yf_csv.close()
              
    except:
        print('******************** Error (420)')


########################################################################

def update_stock_line_item(df2):
    count_symb = 0
    for row_no in range(df2.shape[0]):
        #get share price
        if (df2.iloc[row_no,0]=='$$CASH'):
            df2.iloc[row_no,1]=1
        else:
            df2.iloc[row_no,1]=get_current_stock_price(df2.iloc[row_no,0])
        #calc total_shares
        count_symb+=1
        #print(f'>>>>> ({count_symb}) {df2.iloc[row_no,0]}--->{df2.iloc[row_no,1]}')
        total = 0
        for col_no in range(4,df2.shape[1]):
            total += round(df2.iloc[row_no,col_no],1) 
        df2.iloc[row_no,3] = total

        #calc and save total value
        df2.iloc[row_no,2] = df2.iloc[row_no,1]*df2.iloc[row_no,3]

def main_vg_fd():
    symbs_vg_fd_list = [] # unique entry for symbol 
    matrix_2d = [[0 for x in range(len(col_header_vg_fd))] for y in range(0)] 
    try:

        print('Start reading VG =', datetime.now())
        vanguard_reader(matrix_2d,g_vg_dict,g_data_path+'vg.csv',col_header_vg_fd,symbs_vg_fd_list)
        print('Start reading FD =', datetime.now())

        fidelity_reader(matrix_2d,g_fd_dict,g_data_path+"fd.csv",col_header_vg_fd,symbs_vg_fd_list)

        #print('Start reading FD(2) =', datetime.now())
        #fidelity_reader(matrix_2d,g_fd_mephy_dict,g_data_path+"fd_mephy.csv",col_header_vg_fd,symbs_vg_fd_list)
        #print('End reading FD =', datetime.now())
        
        print('Get stock prices.... =', datetime.now())

        df2 = pd.DataFrame(matrix_2d)

        #update share_price (1), total_value (2) and total_shares (3) columns
        update_stock_line_item(df2)

        df_sorted = df2.sort_values([2], ascending=[False])

        calc_mega_8_holdings(df2)
        
        save2excel(f'{vg_fd_sheet_name}',df_sorted,col_header_vg_fd)  
 
        make_export_for_yf(df_sorted,g_yf_export_vg_fd)



    except:
        print('**** ERROR 264 *****')
        return 0
##########################################################
#tw_reader(matrix_2d,g_vg_dict,g_data_path+'tw.csv',col_header_tw,symbs_tw_list)
def tw_reader(matrix,cvs_file_name,col_header,symb_list):
    fd_last_acc = ''
    try:
        count_0 = 0
        with open(f"{cvs_file_name}") as csv_file:
            print(f"File Opened : [{cvs_file_name}]")
            fd_csv_reader = csv.reader(csv_file, delimiter=',')
            count_0=0
            for row in fd_csv_reader:
                if row[1]=='end':
                    break
                if len(row)==0 or row[0]=='':
                    count_0+=1
                    if (count_0 == 2):
                        break
                    continue
                count_0=0
                tw_acc_name = row[0]
 
                try :             
                    symb = row[1]
                    shares = float(row[3])
                    add_shares(matrix,tw_acc_name,symb,shares,0,col_header,symb_list)   
                         
           
                except ValueError:
                    print('-------line 280', row[0],row[2])
                    print(f'***Fatal Error (242) : symb={symb} Failed to process ({cvs_file_name})') 

    except:
        print(f'***Fatal Error (246) : Unable to process ({cvs_file_name})') 

    return
###########################################################
def main_tw():
    symbs_tw_list = [] # unique entry for symbol 
    matrix_2d = [[0 for x in range(len(col_header_tw))] for y in range(0)] 
    try:
        tw_reader(matrix_2d,g_data_path+'tw.csv',col_header_tw,symbs_tw_list)
        print('TW accounts processing complete =', datetime.now())
    except:
        print(f'*** Error tw_reader failed (512)')
        return False
    
    try:
        df2 = pd.DataFrame(matrix_2d)

        #update share_price (1), total_value (2) and total_shares (3) columns
        update_stock_line_item(df2)

        df_sorted = df2.sort_values([2], ascending=[False])
        
        save2excel(f'{tw_sheet_name}',df_sorted,col_header_tw)  

        #df_sorted = df_sorted.sort_values([2], ascending=[True])

        make_export_for_yf(df_sorted,g_yf_export_tw)

    except:
        print('**** ERROR 474 *****')
        return 0 

###########################################################


if __name__ == "__main__":
    g_init()
    print('starting...')
    main_vg_fd()


    main_tw()
        
    update_summary_sheet()

       #np_array = np.array(data_arry)
    # df2 = pd.DataFrame(np_array,columns=col_header_vg_fd)

    #print(g_matrix)

    #print('shape ',df2.shape)
    
  
    # for i in range(5):
    #     print(df2.iloc[i][1])


    #df2.sort_value(['tot_value'], ascending=[False])


    #print(f'Current Date : {datetime.now()}')

 