#edited from mac_air 2023-07-22 15:17
#Testing 2023-07-14 AM
# edited by Windows Chrome APP 
#Testing Git 2023-07-06
#
#  ***** YTD Tax Info ******
#  1. VG data download range set to YTD
#  2. FD - see X38745588 select ':More' then "Tax Info YTD"

#Last update 2024-05-06  sync to source control and MAC
#2024-07-05 updated from PC
#2024-07-05 MAC data path - /Users/seanleu/code/data/vg_fd_pos 


import csv
from numpy import float32
#from pywin.dialogs.list import test
import xlsxwriter
import pandas as pd
import yfinance as yf # backup for mutual fund
import traceback

import xlwings
import shutil
from datetime import date
from datetime import datetime
from datetime import time

import os
import platform

from mkt_etf_comp import get_etf_components
from mkt_stock_price import get_current_stock_price
from mkt_rd_pos_csv import vanguard_reader
from mkt_rd_pos_csv import fidelity_reader
from mkt_rd_pos_csv import tw_reader2
# install lib
# >pip install requests
# >pip install beautifulsoup4
# >pip install lxml

#from bs4 import BeautifulSoup
#import requests
#import lxml


# OneDrive Folder : '/Users/seanleu/OneDrive/80-股市/00-download/'
# github : https://github.com/ssl1203/vg_fd_stocks.git

if (platform.system()=='Darwin'):
    #old g_data_path = f'{os.path.dirname(__file__)}/../vg_fd_stocks_data/'

    g_code_path = '/Users/seanleu/code/vg_fd/vg_fd_stocks/'
    g_data_path = "/Users/seanleu/code/vg_fd/vg_fd_data/"
    
else:
    g_code_path = 'C:\\Users\\seanl\\##40_Code\\vg_fd\\vg_fd_stocks\\'    
    g_data_path = "C:\\Users\\seanl\\##40_Code\\vg_fd\\vg_fd_data\\"


g_yf_export_vg_fd = f'{g_data_path}yf_export_vg_fd_v2.csv'
g_xl_summary_file = f'{g_data_path}summary-{str(date.today())}_v2.xlsx'
g_yf_export_tw = f'{g_data_path}yf_export_tw_v2.csv'
global g_fd_mephy_download_csv
global g_fd_download_csv
global g_vg_download_csv


def g_init():
    #global g_data_path
    #global g_yf_export_vg_fd
    global g_fd_mephy_download_csv
    global g_fd_download_csv
    global g_vg_download_csv

    if (not os.path.exists(g_data_path)):
        print(f'***** Fatal error, expected fold not exist ({g_data_path})')
        exit()   

    if (os.path.exists(g_xl_summary_file)):
        os.remove(g_xl_summary_file)

    wb = xw.Book()
    #wb.sheets.add('Sheet1')

    try:
        wb.save(g_xl_summary_file)       
    except:
        print(f'**** Not able to save - {g_xl_summary_file}')
        exit(-1)
        
    wb.close()

    #g_yf_export_vg_fd = f'{g_data_path}yf_export_vg_fd.csv'

    # g_fd_mephy_download_csv = f"{g_data_path}fd_mephy.csv"
    # if not os.path.exists(g_fd_mephy_download_csv):
    #     print(f'***** Fatal error, expected file not exist ({g_fd_mephy_download_csv}')
    #     exit()
        

    g_fd_download_csv = f"{g_data_path}fd.csv"
    if not os.path.exists(g_fd_download_csv):
        print(f'***** Fatal error, expected file not exist ({g_fd_download_csv}')
        exit()

    g_vg_download_csv = f"{g_data_path}vg.csv"
    if not os.path.exists(g_vg_download_csv):
        print(f'***** Fatal error, expected file not exist ({g_vg_download_csv}')
        exit()


import xlwings as xw
import numpy as np

import os
#so.getcwd()
import openpyxl


g_row_account = 1
g_row_sum = 2

g_first_data_row = 3

g_first_data_col = 'E'
g_last_data_col = 'E'   # = chr(ord(g_first_data_col) + number_of_accounts - 1)

vg_fd_sheet_name = 'vg+fd'
tw_sheet_name = 'tw'



#si.get_quote_table('NVDA')['PE Ratio (TTM)'] 
#http://theautomatic.net/yahoo_fin-documentation/#stock_info

#  hp_laptop path
#g_working_path = 'C:/Users/seans/OneDrive/80-股市/000-download/2021-07-31/' 


g_vg_dict = {"60965808": "VG_Sean_IRA",
        "58513150":"VG_Sean_ROTH",
        "31141191": "VG_JOIN", 
        "52028985": "VG_Mephy_IRA",
        "15094632":"VG_Mephy_ROTH"}


g_fd_dict = {
        "X38745588" : "FD_JOIN", 
        "168682748":"FD_Sean_IRA",
        "229713633":"FD_Sean_ROTH",
        "32330":"FD_Sean_IRA",
        "225995814" : "FD_Mephy_IRA",
        "233641339":"FD_Mephy_HD"}

# g_fd_mephy_dict = { "225995814" : "FD_Mephy_IRA", 
#         "233641339":"FD_Mephy_HD"}




#symb_dict = {'GOOGL': 'GOOG','SPY':'VOO','VTI':'VOO','FZROX':'VOO'}
#value_dict = {}

col_header_vg_fd = [
    "Symbol",
    "Price",
    "Total Value",
    "Total Shares",
    "VG_Sean_IRA",     
    "VG_Sean_ROTH",
    "VG_JOIN", 
    "VG_Mephy_IRA" ,  
    "VG_Mephy_ROTH" ,

    "FD_JOIN"    ,    
    "FD_Sean_IRA" ,
    "FD_Sean_ROTH" ,  

    "FD_Mephy_IRA", 
    "FD_Mephy_HD",
]

col_header_tw = [
    "Symbol",
    "Price",
    "Total Value",
    "Total Shares",
    "yuanta",     
    "cathay" ,  
    "mega" 
]



#########################################
######### Get ETF Top Holdings ##########
######################################### 
#mega_8 = ['AAPL', 'MSFT', 'AMZN', 'NVDA', 'META', 'TSLA', 'GOOGL', 'GOOG', 'TSM']
#mega8_value_dict = {'AAPL':0, 'MSFT':0, 'AMZN':0, 'NVDA':0, 'META':0, 'TSLA':0, 'GOOGL':0, 'GOOG':0, 'TSM':0}

portfolio_top_holdings = {'NVDA':0, 'TSM':0, 'MSFT':0, 'GOOG':0, 'AMZN':0,'AAPL':0, 'META':0,'AMD':0, 
                'COST':0, 'NFLX':0}

ETF_Values = {'QQQ':0,'VOO':0,'VTI':0,'VUG':0,'SPY':0,'VOOG':0}
mutual_funds_2_ETF = {'FZROX':'VTI', 'VTSAX':'VTI'}




#################################################
def calc_portfolio_top_holdings(df_all_stocks):

    #loop thr all holdings and calc two dictionaries [ETF_Values, portfolio_top_holdings]
    for index, row in df_all_stocks.iterrows():
        symb = row[0]
        value = float(row[2])
        if symb  in portfolio_top_holdings.keys() :
            portfolio_top_holdings[f'{symb}'] += value
        if symb in mutual_funds_2_ETF.keys() :   # convert mutual fund to corresponding ETF
            symb = mutual_funds_2_ETF[f'{symb}'] 
        if symb  in ETF_Values.keys() :
            ETF_Values[f'{symb}'] += value  

    #add ETF_Values component stock values to portfolio_top_holdings 

    for etf in ETF_Values.keys():
        etf_components = get_etf_components(etf)
        for stock in etf_components:
            component_value = etf_components[f'{stock}'] * ETF_Values[f'{etf}'] /100
            if stock == 'GOOGL':
               stock = 'GOOG'

            if stock in portfolio_top_holdings:
                portfolio_top_holdings[f'{stock}'] +=  component_value
        

#################################################

acc_index_dict = {}   # dictionary
pos_index_dict = {}   # dictionary


#pos_list - list of [account_name,symbol,shares]
def add_2_data_matrix(matrix,pos_list,symb_list,col_header):

    # shares can be negative as 'buy' pending transaction 
    for idx,pos in enumerate(pos_list) :   
        try:
            idx_col = col_header.index(pos[0])
        except:
            #idx_col = 4
            #if acc_name == 'yuanta':
            #    print(f'$$$$$$ {acc_name} ==== {col_header[4]}')
            print(f'*** Error : account [{pos[0]}] not found in {col_header_vg_fd}')
            return False

        try:
            #fidility may have duplicate cash positions COREXX and SPAXX
            if pos[1] not in symb_list:
                symb_list.append(pos[1])

                #init all cells in new row to 0's
                new_row = [0 for x in range(len(col_header_vg_fd))]
                new_row[0] = pos[1]
                new_row[1] = 0
                matrix.append(new_row)
            
            idx_row = symb_list.index(pos[1])
            
            #update the total value of a stock in the given accont
            #idx_row = idx of symbol, idx_col = idx of account
            matrix[idx_row][idx_col] = pos[2]+matrix[idx_row][idx_col]  #account
        #     if new_symb == "$$CASH":
        #         price = 1
        #     else :
        #         price = get_current_stock_price(new_symb)
        #     matrix[idx_row][col_header.index("Price")]=price
        except:
            print(f"*** Fatal Error (159) - AddShare {pos[0]} {pos[1]} failed")   




def add_shares(matrix,acc_name,new_symb,shares,price,col_header,symb_list):
    

    # shares can be negative as 'buy' pending transaction     
    try:
        idx_col = col_header.index(acc_name)
    except:
        #idx_col = 4
        #if acc_name == 'yuanta':
        #    print(f'$$$$$$ {acc_name} ==== {col_header[4]}')
        print(f'*** Error : account [{acc_name}] not found in {col_header}')
        return False

    try:
        #fidility may have duplicate cash positions COREXX and SPAXX
        if new_symb not in symb_list:
            symb_list.append(new_symb)

            #init all cells in new row to 0's
            new_row = [0 for x in range(len(col_header))]
            new_row[0] = new_symb
            new_row[1] = price
            matrix.append(new_row)
        
        idx_row = symb_list.index(new_symb)
        
        #update the total value of a stock in the given accont
        #idx_row = idx of symbol, idx_col = idx of account
        matrix[idx_row][idx_col] = shares+matrix[idx_row][idx_col]  #account
    #     if new_symb == "$$CASH":
    #         price = 1
    #     else :
    #         price = get_current_stock_price(new_symb)
    #     matrix[idx_row][col_header.index("Price")]=price
    except:
        print(f"*** Fatal Error (159) - AddShare {acc_name} {new_symb} failed")   

def update_summary_sheet():
    print("Creating summary sheet")

    sheet_name = "Summary"
#########################################
    try :
        wb = xw.Book(g_xl_summary_file)
    except:
        print(f'**** Failed to open {g_xl_summary_file} (302)')
        return

    try:
        wb.sheets.add(sheet_name)
        ws_summary = wb.sheets[sheet_name]
        #ws_summary.range("A1:E25").clear_contains()
    except:
        #wb.sheets.add(sheet_name)
        ws_summary = wb.sheets[sheet_name]

    ws_summary.range('A1:A30').font.bold = True
    ws_summary.range('D1:D30').font.bold = True
    ws_summary.range('E1:E1').font.bold = True

    
    # for border_id in rng(7,13):
    #     rng.Borders(border_id).LineStyle=1
    #     rng.Borders(border_id).Weight=2



####################################
   
    try:
        ws_vg_fd = wb.sheets[vg_fd_sheet_name]
    except:
        print(f'Failed at 404, vg_fd_sheet_name = {vg_fd_sheet_name}')
        return
    
    ws_vg_fd.range('A1:Z1').font.bold = True
    ws_vg_fd.range('A1:A30').font.bold = True


    ws_tw = wb.sheets[tw_sheet_name]

###############################################
    array_2d_vg = pd.DataFrame(ws_vg_fd.range('E1:N2').value)

    ws_summary["A6"].options(pd.DataFrame, header=0, index=False,transpose=True).value = array_2d_vg

    ######### update indices ###########
    #w_sheet.write('B2','S&P-500')
    #w_sheet.write('C2',si.get_live_price('^GSPC'))

    ntd_x_rate = get_current_stock_price("USDTWD=X")
    chf_x_rate = get_current_stock_price("CHFUSD=X")


    ws_summary['B1'].value=f'{date.today()}'
    ws_summary['B2'].value=get_current_stock_price("^GSPC")
    ws_summary['B3'].value=get_current_stock_price("^IXIC")
    ws_summary['B4'].value=get_current_stock_price("NVDA")
    ws_summary['B5'].value=get_current_stock_price("QQQ")



    ws_summary['A1'].value='Date'
    ws_summary['A2'].value='S&P'
    ws_summary['A3'].value='Nasdaq'
    ws_summary['A4'].value='NVDA'
    ws_summary['A5'].value='QQQ'    

    


    ws_summary.range(f'B2:B19').number_format =    "#,##0"
    ws_summary.range(f'E1:E13').number_format =    "#,##0"
    ws_summary.range(f'B20:B21').number_format =   "##.00"

###############################################################
    ws_summary['D1'].value='Total'
    ws_summary['E1'].value='=SUM(B6:B18)'

    ws_summary['D2'].value='VG total'
    ws_summary['E2'].value = '=SUM(B6:B10)'

    ws_summary['D3'].value='Fidelity'
    ws_summary['E3'].value = '=SUM(B11:B15)'

    ws_summary['D4'].value='vg+fd'
    ws_summary['E4'].value = '=E2+E3'

    #ws_summary['D4'].value='FD Mephy'
    #ws_summary['E4'].value='=SUM(B15:B16)+B12'
###################################################
    ws_summary['D6'].value='Sean_IRA'
    ws_summary['E6'].value='=B6+B12'

    ws_summary['D7'].value='Mephy_IRA'
    ws_summary['E7'].value='=B9+B14'

    ws_summary['D8'].value='Sean_Roth'
    ws_summary['E8'].value='=B7+B13'

    ws_summary['D9'].value='Mephy_Roth'
    ws_summary['E9'].value='=B10+B15'

    ws_summary['D10'].value='Saving'
    ws_summary['E10'].value='=B8+B11+B16+B17+B18'
##################################################


    ws_summary['A16'].value = 'QQQ192'
    ws_summary['B16'].value = get_current_stock_price("QQQ")*192

    ws_summary['A17'].value='Taiwan'

    #ws_summary['B17'].value =  ws_tw['C2'].value/31.5
    ws_summary['B17'].value =  float(ws_tw['C2'].value)/ntd_x_rate


    ws_summary['A18'].value='Tecan Stock'
    total_tecn = (get_current_stock_price("TECN.SW")*490-(264*228.6)-(226*236))
    ws_summary['B18'].value = total_tecn*chf_x_rate


    ws_summary['A20'].value='USDTWD=X'
    ws_summary['B20'].value = ntd_x_rate
    ws_summary['A21'].value='CHFUSD=X'
    ws_summary['B21'].value = chf_x_rate

##################################################################
############  Gega_8 #############################################
##############################################################
    
    count=13
    for symb, val in portfolio_top_holdings.items():
        #print(f'OK x, >>>>>  {symb} = {val}')

        #stock symbol
        ws_summary[f'D{count}'].value = symb

        #stock value
        ws_summary[f'E{count}'].value = val

        #value percentage of total portfolio
        ws_summary[f'F{count}'].value = f'=E{count}/E1'
        count=count+1
   

    #ws_summary.range(f'E13:E30').number_format =   "##.00"
    ws_summary.range(f'E13:E30').number_format =   "$#,##0"
    ws_summary.range(f'F13:F30').number_format =   '0.00%'


    print('----------- Summary Sheet Complete ---------------')
    #ws_summary['E6'].formula = tecn_total * ws_summary['E16'].value
   







########################################################################
def save2excel(sheet_name,df_sorted,col_header):

    first_acc_col = 'E'
    last_acc_col = chr(ord(first_acc_col)+df_sorted.shape[1]-4-1)  # 4 is the none account columns (overhead columns)
    first_stock_row = 3
    last_stock_row = first_stock_row+df_sorted.shape[0]-1

    try :     
        wb = xw.Book(g_xl_summary_file)
    except:
        print(f'**** Failed to open {g_xl_summary_file}')
        return
    
######################################
# Cleanup vg_fd by delete it and create a fresh one
    try:
        if (not sheet_name in wb.sheet_names):
            wb.sheets.add(sheet_name)
    except:
        print('**** ERROR 458 *****', sheet_name)
####################################
    ws = wb.sheets[f'{sheet_name}']

    #df = pd.DataFrame(data_arry)
    ws[f"A1"].options(pd.DataFrame, header=0, index=False, transpose=True).value = pd.DataFrame(col_header)
    ws[f"A{first_stock_row}"].options(pd.DataFrame, header=0, index=False, expand='table').value = df_sorted

    ws[f'C2'].formula=f'=SUM(C{first_stock_row}:C{last_stock_row})' 

    for col_id in range(ord(first_acc_col), ord(last_acc_col)+1):
        ws[f'{chr(col_id)}2'].formula = f'=SUMPRODUCT(B{first_stock_row}:B{last_stock_row},{chr(col_id)}{first_stock_row}:{chr(col_id)}{last_stock_row})' 

    ws['A2'].formula = f'=TEXT(TODAY(), "yyyy-mm-dd")'
    ws['B2'].value = datetime.now().strftime("%X")

    
    ws.range(f'B3:{last_acc_col}{last_stock_row}').number_format =    "#,##0"
    ws.range(f'C2:{last_acc_col}2').number_format =    "#,##0"
    ws.range('C2:C2').font.bold = True
    #ws_vg_fd.range(f'C:{g_row_sum}').number_format = "#,##0.#000"
    #print(f'======OK y, ========== Total Value :',  "{:,.0f}".format(ws[f'C2'].value), ' ===================')

    return True


########################################################################
def make_export_for_yf(df_holdings,yf_export_file):
    title = 'Symbol,Current Price,Date,Time,Change,Open,High,Low,Volume,Trade Date,Purchase Price,Quantity,Commission,High Limit,Low Limit,Comment\r'
    
    try:
        with open(yf_export_file, 'w') as yf_csv:
            yf_csv.write(title)
            # reverse order from small to large position
            for stock_position in df_holdings.iloc[::-1].itertuples(): 
                yf_csv.write(f'{stock_position[1]},{stock_position[2]},,,,,,,,,,{stock_position[4]},,,,\r') 
            yf_csv.close()
              
    except Exception as e:
        print(e)
        traceback.print_exc()
        print('******************** Error (420)')


########################################################################
## #update share_price (1), total_value (2) and total_shares (3) columns
#########################################################################

def update_stock_line_item(df2):
    count_symb = 0
    for row_no in range(df2.shape[0]):
        #get share price
        if (df2.iloc[row_no,0]=='$$CASH'):
            df2.iloc[row_no,1]=1
        else:
            df2.iloc[row_no,1]=get_current_stock_price(df2.iloc[row_no,0])
        #calc total_shares
        count_symb+=1
        #print(f'>>>>> ({count_symb}) {df2.iloc[row_no,0]}--->{df2.iloc[row_no,1]}')
        total = 0
        for col_no in range(4,df2.shape[1]):
            total += round(df2.iloc[row_no,col_no],1) 
        df2.iloc[row_no,3] = total

        #calc and save total value
        df2.iloc[row_no,2] = df2.iloc[row_no,1]*df2.iloc[row_no,3]



        if df2.iloc[row_no,0] == '2330.tw':
            val_2330 = df2.iloc[row_no,2]/get_current_stock_price("USDTWD=X")
            portfolio_top_holdings['TSM']+=val_2330
            #print('OK 6, Found TSM ---',val_2330)
            

############################################################################
# Read VG+FD data into following 2-D array then converted to DataFrame, 
# calcuate empty fields (such as Total shares, tatal value...) 
# before saving into Excel
#---------------------------------------------------------------------------
# {Symbol,	     Price,	        Total Value,	Total Shares,	VG_Sean_IRA,	
# VG_Sean_ROTH,	 VG_JOIN,	    VG_Mephy_IRA,	VG_Mephy_ROTH,	FD_JOIN,	
# FD_Sean_IRA,	 FD_Sean_ROTH,	FD_Mephy_IRA,	FD_Mephy_HD}
############################################################################
acc_pos_3d = [[0 for x in range(10)] for y in range(47)]
symb_list = []
acc_list = []
def main_vg_fd():
    symbs_vg_fd_list = [] # unique entry for symbol 
    matrix_2d = [[0 for x in range(len(col_header_vg_fd))] for y in range(0)] 
    
    try:

        print('Start reading VG =', datetime.now())
        vg_list = vanguard_reader(g_vg_dict,g_data_path+'vg.csv')       
        add_2_data_matrix(matrix_2d,vg_list,symbs_vg_fd_list,col_header_vg_fd)
        
        print('Start reading FD =', datetime.now())
        fd_list = fidelity_reader(g_fd_dict,g_data_path+"fd.csv")
        add_2_data_matrix(matrix_2d,fd_list,symbs_vg_fd_list,col_header_vg_fd)


        print('Get stock prices.... =', datetime.now())

        df2 = pd.DataFrame(matrix_2d)

        #set all columns except symbol column in df2 to float type
        rows, columns = df2.shape
        for x in range(1,columns):
            df2[[x]] = df2[[x]].astype(float)

        #update share_price (1), total_value (2) and total_shares (3) columns
        update_stock_line_item(df2)

        #print("OK 1")

        df_sorted = df2.sort_values([2], ascending=[False])

        #print("OK 2")

        #calc_mega_8_holdings(df2)  sean
        calc_portfolio_top_holdings(df2)



        #print("OK 3")
        
        save2excel(f'{vg_fd_sheet_name}',df_sorted,col_header_vg_fd)  

        #print("OK 4")
 
        make_export_for_yf(df_sorted,g_yf_export_vg_fd)

        #print("OK 6")

    except Exception as e:
        print(e)
        traceback.print_exc()
        print('**** ERROR 264 *****')
        return 0
def main_tw():
    symbs_tw_list = [] # unique entry for symbol 
    matrix_2d = [[0 for x in range(len(col_header_tw))] for y in range(0)]
    
    try:
        tw_list = tw_reader2(g_data_path+'tw2.csv')
        add_2_data_matrix(matrix_2d,tw_list,symbs_tw_list,col_header_tw)

    
        #print('OK z, TW accounts processing complete =', datetime.now())
    except:
        print(f'*** Error tw_reader failed (512)')
        return False
    
    try:
        df2 = pd.DataFrame(matrix_2d)

        #update share_price (1), total_value (2) and total_shares (3) columns
        update_stock_line_item(df2)

        df_sorted = df2.sort_values([2], ascending=[False])
        
        save2excel(f'{tw_sheet_name}',df_sorted,col_header_tw)  

        #df_sorted = df_sorted.sort_values([2], ascending=[True])

        make_export_for_yf(df_sorted,g_yf_export_tw)

    except  Exception as e:
        print(e)
        traceback.print_exc()
        print('**** ERROR 474 *****')
        return 0 

###########################################################


if __name__ == "__main__":
    g_init()
    print('starting...')
    main_vg_fd()


    main_tw()
        
    update_summary_sheet()

       #np_array = np.array(data_arry)
    # df2 = pd.DataFrame(np_array,columns=col_header_vg_fd)

    #print(g_matrix)

    #print('shape ',df2.shape)
    
  
    # for i in range(5):
    #     print(df2.iloc[i][1])


    #df2.sort_value(['tot_value'], ascending=[False])


    #print(f'Current Date : {datetime.now()}')

 